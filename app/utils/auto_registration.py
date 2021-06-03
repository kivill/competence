from app.models.profile import Profile
from openpyxl import load_workbook
from app.utils.utils import validate_phone
from app.utils.utils import generate_key
import os
import uuid
import datetime
import logging
import threading
from app import builder
from app.utils.utils import debug
from app.models.delivery import Delivery
from app.utils.notifications.sender import NotificationSender


class AutoRegistration:
    def __init__(self, filename, utm_source, delivery, text_template=""):
        self.filename = filename
        self.utm_source = utm_source
        self.text_template = text_template
        self.delivery = delivery

    def registrate(self):
        wb = load_workbook('./tmp/'+self.filename, data_only=True)
        anotherSheet = wb.active
        start_time = datetime.datetime.now().timestamp()
        for row in anotherSheet.values:
            if row[0]:
                if validate_phone(int(row[0])):
                    try:
                        user = Profile.create(email=str(int(row[0]))+'@mail.ru',
                                        password='default_password',
                                        name='',
                                        phone=int(row[0]),
                                        uuid=str(uuid.uuid4()),
                                        utm_source=self.utm_source,
                                        is_phone_confirm='true',
                                        api_key=generate_key())

                        #TODO check event firing with empty response object
                        builder.event.fire('user:account:created', user, request={})
                    except Exception as e:
                        logging.exception('Cant create profile: {e}'.format(e=e))
                    # Create sms and .email notification chains
                    try:
                        from app.classes.task import TaskPlanner
                        TaskPlanner.create('sms', user)
                        TaskPlanner.create('call', user)
                    except Exception as e:
                        logging.exception('Failed to create task: {e}'.format(e=e))
        stop_time = datetime.datetime.now().timestamp()
        logging.info('AutoReg complete in {time}'.format(time=stop_time-start_time))
        os.remove('./tmp/'+self.filename)

    def send_auto_registration_sms(self):
        wb = load_workbook('./tmp/'+self.filename, data_only=True)
        anotherSheet = wb.active
        users = []
        for row in anotherSheet.values:
            if row[0]:
                if validate_phone(int(row[0])):
                    users.append({'phone': int(row[0])})
        settings = {}
        settings['text_template'] = self.text_template
        settings['utm_source'] = self.utm_source

        try:
            sender = NotificationSender(type='sms', settings=settings, delivery=self.delivery, users=users)
            thread = threading.Thread(name='NotificationSender',
                                      target=sender.send_auto_registration_sms)
            thread.start()
        except Exception as e:
            logging.exception('Cant create sms (auto_registration): {e}'.format(e=e))

        os.remove('./tmp/'+self.filename)


    def send_auto_registration_call(self):
        wb = load_workbook('./tmp/'+self.filename, data_only=True)
        anotherSheet = wb.active
        users = []
        for row in anotherSheet.values:
            if row[0]:
                if validate_phone(int(row[0])):
                    users.append({'phone': int(row[0])})
        settings = {}
        settings['utm_source'] = self.utm_source
        try:
            sender = NotificationSender(type='call', settings=settings, delivery=self.delivery, users=users)
            thread = threading.Thread(name='NotificationSender',
                                      target=sender.send_auto_registration_call)
            thread.start()
        except Exception as e:
            logging.exception('Cant create call (auto_registration): {e}'.format(e=e))

        os.remove('./tmp/'+self.filename)